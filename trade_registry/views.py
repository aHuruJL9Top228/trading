from django.db.models import Count, Sum, Avg
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_GET
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.workbook import Workbook

from .forms import TradingObjectForm, PersonForm, OrganizationForm, SquareRozForm, SquareOptForm
from .models import Object, Person, SpeciesObject, ObjectOrganization, TradingMark, Municipality, ActivityKind, \
    ObjectType, FormProperty, Organization, KindObject, SquareRoz, SquareOpt


def main(request):
    return render(request, 'trade_registry/main.html')


def trade(request):
    # Получаем все записи из модели LegalEntity
    objects = Object.objects.filter(species_object_id=1)
    # Получаем параметры фильтрации из GET-запроса
    object_name = request.GET.get('object_name')
    kod_okpo = request.GET.get('kod_okpo')
    address = request.GET.get('address')
    trading_mark = request.GET.get('trading_mark')
    mo = request.GET.get('mo')
    type_object = request.GET.get('type_object')
    kind_activity = request.GET.get('kind_activity')
    form_property = request.GET.get('form_property')


    # Фильтрация по полям
    if object_name:
        objects = objects.filter(object_name__icontains=object_name)

    if kod_okpo:
        objects = objects.filter(kod_okpo__icontains=kod_okpo)

    if address:
        objects = objects.filter(address__icontains=address)

    if trading_mark:
        try:
            objects = objects.filter(trading_mark_id=int(trading_mark))
        except ValueError:
            pass  # Если не число — игнорируем

    if mo:
        try:
            objects = objects.filter(mo_id=int(mo))
        except ValueError:
            pass

    if type_object:
        try:
            objects = objects.filter(type_object_id=int(type_object))
        except ValueError:
            pass

    if kind_activity:
        try:
            objects = objects.filter(kind_activity_id=int(kind_activity))
        except ValueError:
            pass

    if form_property:
        try:
            objects = objects.filter(form_property_id=int(form_property))
        except ValueError:
            pass

    print(TradingMark.objects.all())

    # Передаём в контекст также значения фильтров (чтобы они сохранялись в полях)
    context = {
        'objects': objects,
        'object_name': object_name or '',
        'kod_okpo': kod_okpo or '',
        'address': address or '',
        'trading_mark': trading_mark or '',
        'mo': mo or '',
        'type_object': type_object or '',
        'kind_activity': kind_activity or '',
        'form_property': form_property or '',
        # Передаём списки для выпадающих списков
        'trading_marks': TradingMark.objects.all(), # бренд
        'municipalities': Municipality.objects.all(), # мо
        'object_types': ObjectType.objects.all(), # прод/непрод
        'activity_kinds': ActivityKind.objects.all(), # мясо рыба мебель
        'form_properties': FormProperty.objects.all(), # форма собственности

    }
    return render(request, 'trade_registry/trade.html', context)


def meal(request):
    # Получаем все записи из модели LegalEntity
    objects = Object.objects.all()
    context = {
        'objects': objects,
    }
    return render(request, 'trade_registry/meal.html', context)


def services(request):
    # Получаем все записи из модели LegalEntity
    objects = Object.objects.all()
    context = {
        'objects': objects,
    }
    return render(request, 'trade_registry/services.html', context)


def add_trade(request):
    speciesObject = SpeciesObject.objects.get(name='Торговля')

    # Получаем ID для розничной и оптовой разновидностей
    try:
        retail_kind = KindObject.objects.get(name="Розничная")
        wholesale_kind = KindObject.objects.get(name="Оптовая")
    except KindObject.DoesNotExist:
        # Если объекты не найдены, создаем их
        retail_kind = KindObject.objects.create(name="Розничная")
        wholesale_kind = KindObject.objects.create(name="Оптовая")

    # Инициализируем формы
    square_roz_form = SquareRozForm()
    square_opt_form = SquareOptForm()

    if request.method == 'POST':
        form = TradingObjectForm(request.POST)

        # Проверяем, какую форму площадей нужно обработать
        kind_of_object_id = request.POST.get('kind_of_object')

        # Определяем тип объекта
        is_retail = kind_of_object_id == str(retail_kind.id)
        is_wholesale = kind_of_object_id == str(wholesale_kind.id)

        # Создаем соответствующие формы
        square_roz_form = SquareRozForm(request.POST if is_retail else None)
        square_opt_form = SquareOptForm(request.POST if is_wholesale else None)

        # Проверяем валидность основной формы
        if form.is_valid():
            obj = form.save(commit=False)
            obj.species_object = speciesObject  # Программно задаём вид объекта

            # Сохраняем основной объект
            obj.save()

            # Обрабатываем дополнительные формы
            if is_retail:
                if square_roz_form.is_valid():
                    square_roz = square_roz_form.save(commit=False)
                    square_roz.object = obj
                    square_roz.save()
                else:
                    # Если форма площадей не валидна, удаляем объект
                    obj.delete()
                    # И возвращаем форму с ошибками
                    return render(request, 'trade_registry/add_trade.html', {
                        'form': form,
                        'square_roz_form': square_roz_form,
                        'square_opt_form': square_opt_form,
                        'retail_kind_id': retail_kind.id,
                        'wholesale_kind_id': wholesale_kind.id
                    })
            elif is_wholesale:
                if square_opt_form.is_valid():
                    square_opt = square_opt_form.save(commit=False)
                    square_opt.object = obj
                    square_opt.save()
                else:
                    # Если форма площадей не валидна, удаляем объект
                    obj.delete()
                    # И возвращаем форму с ошибками
                    return render(request, 'trade_registry/add_trade.html', {
                        'form': form,
                        'square_roz_form': square_roz_form,
                        'square_opt_form': square_opt_form,
                        'retail_kind_id': retail_kind.id,
                        'wholesale_kind_id': wholesale_kind.id
                    })

            return redirect('trade')
    else:
        form = TradingObjectForm()

    return render(request, 'trade_registry/add_trade.html', {
        'form': form,
        'square_roz_form': square_roz_form,
        'square_opt_form': square_opt_form,
        'retail_kind_id': retail_kind.id,
        'wholesale_kind_id': wholesale_kind.id
    })


def add_person(request):
    if request.method == 'POST':
        form = PersonForm(request.POST)
        if form.is_valid():
            form.save()
            # После сохранения можно перенаправить обратно на страницу trade
            return redirect('trade')
    else:
        form = PersonForm()

    return render(request, 'trade_registry/add_person.html', {'form': form})


def add_organization(request):
    if request.method == 'POST':
        form = OrganizationForm(request.POST)
        if form.is_valid():
            # Сохраняем организацию
            organization = form.save()

            # Получаем список строковых ID и конвертируем в int
            linked_object_ids = request.POST.getlist('linked_objects')
            linked_object_ids = [int(obj_id) for obj_id in linked_object_ids if obj_id.isdigit()]

            print("Organization ID:", organization.id)
            print("Organization name:", organization.name)
            print("Linked object IDs:", linked_object_ids)

            # Создаём связи в таблице ObjectOrganization
            created_links = []
            for obj_id in linked_object_ids:
                obj = Object.objects.get(id=obj_id)  # Убедимся, что объект существует
                link, created = ObjectOrganization.objects.get_or_create(
                    object=obj,
                    organization=organization
                )
                created_links.append(link)

            # Можно подсчитать, сколько создано
            print(f"Создано связей: {len(created_links)}")
            return redirect('trade')
    else:
        form = OrganizationForm()

    return render(request, 'trade_registry/add_organization.html', {'form': form})


@require_GET
def object_search(request):
    q = request.GET.get('q', '')
    print(q)
    if len(q) < 2:
        return JsonResponse([], safe=False)

    objects = Object.objects.filter(object_name__icontains=q)[:10]
    results = [
        {'id': obj.id, 'object_name': obj.object_name}
        for obj in objects
    ]
    print(results)
    return JsonResponse(results, safe=False)


def view_object(request, object_id):
    object_item = get_object_or_404(Object, id=object_id)

    # Определяем типы объектов
    try:
        retail_kind = KindObject.objects.get(name="Розничная")
        wholesale_kind = KindObject.objects.get(name="Оптовая")
    except KindObject.DoesNotExist:
        retail_kind = None
        wholesale_kind = None

    # Создаем форму для основного объекта
    form = TradingObjectForm(instance=object_item)

    # Делаем все поля только для чтения
    for field in form.fields:
        form.fields[field].widget.attrs['readonly'] = True
        form.fields[field].widget.attrs['disabled'] = True

    # Получаем соответствующую форму площади
    square_roz_form = None
    square_opt_form = None

    # Проверяем тип объекта и получаем данные о площади
    if retail_kind and object_item.kind_of_object == retail_kind:
        square_roz = SquareRoz.objects.filter(object=object_item).first()
        if square_roz:
            square_roz_form = SquareRozForm(instance=square_roz)
            # Делаем поля только для чтения
            for field in square_roz_form.fields:
                square_roz_form.fields[field].widget.attrs['readonly'] = True
                square_roz_form.fields[field].widget.attrs['disabled'] = True
    elif wholesale_kind and object_item.kind_of_object == wholesale_kind:
        square_opt = SquareOpt.objects.filter(object=object_item).first()
        if square_opt:
            square_opt_form = SquareOptForm(instance=square_opt)
            # Делаем поля только для чтения
            for field in square_opt_form.fields:
                square_opt_form.fields[field].widget.attrs['readonly'] = True
                square_opt_form.fields[field].widget.attrs['disabled'] = True

    return render(request, 'trade_registry/view_object.html', {
        'form': form,
        'object': object_item,
        'square_roz_form': square_roz_form,
        'square_opt_form': square_opt_form,
        'retail_kind_id': retail_kind.id if retail_kind else None,
        'wholesale_kind_id': wholesale_kind.id if wholesale_kind else None
    })


def edit_object(request, object_id):
    object_item = get_object_or_404(Object, id=object_id)

    # Сохраняем текущий тип объекта ДО изменений
    old_kind_of_object = object_item.kind_of_object

    # Определяем типы объектов
    try:
        retail_kind = KindObject.objects.get(name="Розничная")
        wholesale_kind = KindObject.objects.get(name="Оптовая")
    except KindObject.DoesNotExist:
        retail_kind = None
        wholesale_kind = None

    # Переменные для форм
    square_roz_form = None
    square_opt_form = None
    show_square_roz = False
    show_square_opt = False

    if request.method == 'POST':
        # Создаем форму с POST-данными
        form = TradingObjectForm(request.POST, instance=object_item)

        # Получаем новый тип объекта из POST-данных
        new_kind_of_object_id = request.POST.get('kind_of_object')

        # Определяем, изменился ли тип объекта
        kind_changed = False
        if old_kind_of_object and new_kind_of_object_id:
            kind_changed = str(old_kind_of_object.id) != new_kind_of_object_id
        elif (old_kind_of_object and not new_kind_of_object_id) or (not old_kind_of_object and new_kind_of_object_id):
            kind_changed = True

        # Обрабатываем форму площади в зависимости от типа объекта
        if retail_kind and new_kind_of_object_id == str(retail_kind.id):
            # Если тип изменился (включая случай, когда был NULL)
            if kind_changed:
                # Удаляем старую запись, если она существует
                SquareOpt.objects.filter(object=object_item).delete()
                # Создаем НОВУЮ форму для розницы (без instance)
                square_roz_form = SquareRozForm(request.POST)
            else:
                # Тип не изменился, используем существующую запись (если есть)
                square_roz = SquareRoz.objects.filter(object=object_item).first()
                square_roz_form = SquareRozForm(request.POST, instance=square_roz)

            show_square_roz = True
        elif wholesale_kind and new_kind_of_object_id == str(wholesale_kind.id):
            # Если тип изменился (включая случай, когда был NULL)
            if kind_changed:
                # Удаляем старую запись, если она существует
                SquareRoz.objects.filter(object=object_item).delete()
                # Создаем НОВУЮ форму для опта (без instance)
                square_opt_form = SquareOptForm(request.POST)
            else:
                # Тип не изменился, используем существующую запись (если есть)
                square_opt = SquareOpt.objects.filter(object=object_item).first()
                square_opt_form = SquareOptForm(request.POST, instance=square_opt)

            show_square_opt = True

        if form.is_valid():
            obj = form.save()

            # Сохраняем форму площади, если она валидна
            if square_roz_form and square_roz_form.is_valid():
                square_roz = square_roz_form.save(commit=False)
                square_roz.object = obj
                square_roz.save()
            elif square_opt_form and square_opt_form.is_valid():
                square_opt = square_opt_form.save(commit=False)
                square_opt.object = obj
                square_opt.save()

            return redirect('trade')
    else:
        form = TradingObjectForm(instance=object_item)

        # Определяем, какую форму площади показывать
        if retail_kind and object_item.kind_of_object == retail_kind:
            # Для розничного объекта показываем форму с данными (если есть)
            square_roz = SquareRoz.objects.filter(object=object_item).first()
            square_roz_form = SquareRozForm(instance=square_roz)
            # ВСЕГДА создаем пустую форму для опта (но не показываем её)
            square_opt_form = SquareOptForm()
            show_square_roz = True
        elif wholesale_kind and object_item.kind_of_object == wholesale_kind:
            # Для оптового объекта показываем форму с данными (если есть)
            square_opt = SquareOpt.objects.filter(object=object_item).first()
            square_opt_form = SquareOptForm(instance=square_opt)
            # ВСЕГДА создаем пустую форму для розницы (но не показываем её)
            square_roz_form = SquareRozForm()
            show_square_opt = True
        else:
            # Если тип объекта не установлен или неизвестен
            square_roz_form = SquareRozForm()
            square_opt_form = SquareOptForm()

    return render(request, 'trade_registry/edit_object.html', {
        'form': form,
        'object': object_item,
        'square_roz_form': square_roz_form,
        'square_opt_form': square_opt_form,
        'show_square_roz': show_square_roz,
        'show_square_opt': show_square_opt,
        'retail_kind_id': retail_kind.id if retail_kind else None,
        'wholesale_kind_id': wholesale_kind.id if wholesale_kind else None
    })


def owners(request):
    # Получаем всех Person, которые являются owner организаций,
    # связанных с объектами, у которых species_object.id = 1
    owners = Person.objects.filter(
        organization__objectorganization__object__species_object__id=1
    ).distinct()

    return render(request, 'trade_registry/owners.html', {'owners': owners})


def organization(request):
    # Получаем все организации, связанные с объектами, у которых species_object.id = 1
    organizations = Organization.objects.filter(
        objectorganization__object__species_object__id=1
    ).distinct()

    return render(request, 'trade_registry/organization.html', {'organizations': organizations})


def report(request):
    # Получаем все записи из модели LegalEntity
    objects = Object.objects.all()
    context = {
        'objects': objects,
    }
    return render(request, 'trade_registry/report.html', context)


from django.db.models import Count, Avg, Sum, Q


def export_trade_marks_report(request):
    # Получаем торговые марки с количеством ТОРГОВЫХ объектов больше 10
    trade_marks = TradingMark.objects.annotate(
        object_count=Count('object', filter=Q(object__species_object_id=1))
    ).filter(object_count__gt=10).order_by('-object_count')

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по торговым маркам"

    # Стили
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    total_font = Font(bold=True)

    # Заголовок
    ws.merge_cells('A1:G1')
    ws['A1'] = "Приложение 10"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:G2')
    ws['A2'] = "от « » июля 2025 года №"
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A4:G4')
    ws[
        'A4'] = "Обобщенные сведения, содержащиеся в торговом реестре Белгородской области, о торговых марках (брендов) с количеством торговых объектов больше 10 единиц"
    ws['A4'].font = Font(bold=True)
    ws['A4'].alignment = center_alignment
    ws.row_dimensions[4].height = 40

    # Шапка таблицы
    headers = [
        "Торговая марка (бренд)",
        "Количество торговых объектов (ед.)",
        "Средняя численность работников хозяйствующих субъектов (чел.)",
        "Площадь (кв.м.)",
        "",
        "торгового объекта",
        ""
    ]

    subheaders = [
        "",
        "",
        "",
        "общая",
        "иное законное основание в т.ч. аренда",
        "на праве собственности",
        "иное законное основание в т.ч. аренда"
    ]

    # Записываем шапку
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border_style

    # Записываем подзаголовки
    for col, subheader in enumerate(subheaders, 1):
        cell = ws.cell(row=7, column=col, value=subheader)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border_style
        if subheader:
            cell.fill = gray_fill

    # Объединяем ячейки для "Площадь (кв.м.)"
    ws.merge_cells('D6:E6')
    ws['D6'].alignment = center_alignment

    # Объединяем ячейки для "торгового объекта"
    ws.merge_cells('F6:G6')
    ws['F6'].alignment = center_alignment

    # Устанавливаем ширину колонок
    column_widths = [20, 25, 40, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Инициализируем переменные для суммирования
    total_objects = 0
    total_avg_workers = 0
    total_owner = 0
    total_law = 0
    total_trade_owner = 0
    total_trade_law = 0

    # Заполняем данными
    row = 8
    for mark in trade_marks:
        # Получаем ТОЛЬКО ТОРГОВЫЕ объекты для этой торговой марки
        objects = Object.objects.filter(trading_mark=mark, species_object_id=1)

        # Рассчитываем площади ТОЛЬКО для розничных ТОРГОВЫХ объектов
        roznica = SquareRoz.objects.filter(object__trading_mark=mark, object__species_object_id=1)
        total_owner_roz = roznica.aggregate(Sum('all_owner'))['all_owner__sum'] or 0
        total_law_roz = roznica.aggregate(Sum('all_law'))['all_law__sum'] or 0
        trade_owner_roz = roznica.aggregate(Sum('trade_obj_owner'))['trade_obj_owner__sum'] or 0
        trade_law_roz = roznica.aggregate(Sum('trade_obj_law'))['trade_obj_law__sum'] or 0

        # Рассчитываем площади ТОЛЬКО для оптовых ТОРГОВЫХ объектов
        opt = SquareOpt.objects.filter(object__trading_mark=mark, object__species_object_id=1)
        total_owner_opt = opt.aggregate(Sum('warehouse_square'))['warehouse_square__sum'] or 0
        # Для опта нет такого разделения в вашей модели
        total_law_opt = 0
        trade_owner_opt = 0
        trade_law_opt = 0

        # Общие площади
        total_owner_mark = total_owner_roz + total_owner_opt
        total_law_mark = total_law_roz + total_law_opt
        trade_owner_mark = trade_owner_roz + trade_owner_opt
        trade_law_mark = trade_law_roz + trade_law_opt

        # Средняя численность работников ТОЛЬКО для торговых объектов
        avg_workers = objects.aggregate(Avg('avg_workers'))['avg_workers__avg'] or 0

        # Записываем данные
        ws.cell(row=row, column=1, value=mark.brand_name.upper())
        ws.cell(row=row, column=2, value=objects.count())
        ws.cell(row=row, column=3, value=round(avg_workers, 1))
        ws.cell(row=row, column=4, value=round(total_owner_mark, 3))
        ws.cell(row=row, column=5, value=round(total_law_mark, 3))
        ws.cell(row=row, column=6, value=round(trade_owner_mark, 3))
        ws.cell(row=row, column=7, value=round(trade_law_mark, 3))

        # Применяем стили
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = border_style
            if col > 3:  # Для числовых значений
                cell.number_format = '#,##0.000'

        # Накапливаем суммы для строки "Итого"
        total_objects += objects.count()
        total_avg_workers += avg_workers
        total_owner += total_owner_mark
        total_law += total_law_mark
        total_trade_owner += trade_owner_mark
        total_trade_law += trade_law_mark

        row += 1

    # Добавляем строку "Итого"
    ws.cell(row=row, column=1, value="Итого")
    ws.cell(row=row, column=2, value=total_objects)
    ws.cell(row=row, column=3, value=round(total_avg_workers, 1))
    ws.cell(row=row, column=4, value=round(total_owner, 3))
    ws.cell(row=row, column=5, value=round(total_law, 3))
    ws.cell(row=row, column=6, value=round(total_trade_owner, 3))
    ws.cell(row=row, column=7, value=round(total_trade_law, 3))

    # Применяем стили для строки "Итого"
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
    cell.font = total_font
    cell.border = border_style
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    if col > 1:  # Для числовых значений
        cell.number_format = '#,##0.000' if col > 3 else '#,##0.0'

    row += 1  # Переходим к следующей строке

    # Номера строк
    for col in range(1, 8):
        ws.cell(row=row, column=col, value=col)
    ws.cell(row=row, column=col).font = header_font
    ws.cell(row=row, column=col).border = border_style

    # Устанавливаем заголовки страницы для печати
    ws.oddHeader.left.text = "Приложение 10"
    ws.oddHeader.right.text = "от « » июля 2025 года №"

    # Сохраняем в поток
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=otchet_po_torgovym_markam.xlsx'
    wb.save(response)

    return response